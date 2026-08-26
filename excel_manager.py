# -*- coding: utf-8 -*-
"""
Handles all Excel bookkeeping for the coffee shop POS app.

Design (per day sheet, named with the Gregorian date, e.g. 2026-08-25):

  1) A SUMMARY block at the top: one row per menu item, with the sold
     count and total amount computed by COUNTIF/SUMIF formulas that read
     from the transaction log below. This lets prices differ from sale to
     sale (discounts, VIP prices, etc.) while the summary still totals
     correctly, and it recalculates live whenever the file is opened in
     Excel.

  2) A LOG block below it: one row per individual sale (time, item, the
     exact price charged for that sale). This is the source of truth.

Because openpyxl cannot evaluate formulas itself, the Python side always
computes "today's count" by counting/scanning the log rows directly -
the formulas in the sheet are for when a human opens the file in Excel.
"""

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from menu_data import MENU_ITEMS
import date_utils

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sales_data")
WORKBOOK_PATH = os.path.join(DATA_DIR, "sales.xlsx")

N = len(MENU_ITEMS)

# --- Summary block layout ---
SUMMARY_TITLE_ROW = 1
SUMMARY_HEADER_ROW = 2
SUMMARY_FIRST_ROW = 3
SUMMARY_LAST_ROW = SUMMARY_FIRST_ROW + N - 1          # one row per item
TOTAL_ROW = SUMMARY_LAST_ROW + 1
S_COL_KEY, S_COL_NAME, S_COL_COUNT, S_COL_TOTAL = 1, 2, 3, 4

# --- Log block layout ---
LOG_TITLE_ROW = TOTAL_ROW + 2
LOG_HEADER_ROW = LOG_TITLE_ROW + 1
LOG_FIRST_ROW = LOG_HEADER_ROW + 1                    # first possible data row
L_COL_NO, L_COL_TIME, L_COL_KEY, L_COL_NAME, L_COL_PRICE = 1, 2, 3, 4, 5

LOG_KEY_RANGE = f"C{LOG_FIRST_ROW}:C100000"
LOG_PRICE_RANGE = f"E{LOG_FIRST_ROW}:E100000"

HEADER_FILL = PatternFill(start_color="4E342E", end_color="4E342E", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
TOTAL_FILL = PatternFill(start_color="D7CCC8", end_color="D7CCC8", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=13, color="4E342E")
THIN_BORDER = Border(*([Side(style="thin", color="BFBFBF")] * 4))
CENTER = Alignment(horizontal="center", vertical="center")


def _today_str():
    return date_utils.today_str()


def _ensure_dir():
    os.makedirs(DATA_DIR, exist_ok=True)


def _new_sheet(wb, sheet_name):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.rightToLeft = True

    ws.cell(row=SUMMARY_TITLE_ROW, column=1, value=f"خلاصه فروش روز {sheet_name}").font = TITLE_FONT

    headers = ["کلید", "نام نوشیدنی", "تعداد فروش", "مبلغ کل (تومان)"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=SUMMARY_HEADER_ROW, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for i, item in enumerate(MENU_ITEMS):
        r = SUMMARY_FIRST_ROW + i
        ws.cell(row=r, column=S_COL_KEY, value=item["key"]).border = THIN_BORDER
        ws.cell(row=r, column=S_COL_NAME, value=item["name"]).border = THIN_BORDER
        count_cell = ws.cell(row=r, column=S_COL_COUNT,
                              value=f"=COUNTIF({LOG_KEY_RANGE},A{r})")
        total_cell = ws.cell(row=r, column=S_COL_TOTAL,
                              value=f"=SUMIF({LOG_KEY_RANGE},A{r},{LOG_PRICE_RANGE})")
        count_cell.border = THIN_BORDER
        total_cell.border = THIN_BORDER
        count_cell.alignment = CENTER
        total_cell.alignment = CENTER

    ws.cell(row=TOTAL_ROW, column=S_COL_NAME, value="مجموع کل")
    ws.cell(row=TOTAL_ROW, column=S_COL_COUNT,
            value=f"=SUM(C{SUMMARY_FIRST_ROW}:C{SUMMARY_LAST_ROW})")
    ws.cell(row=TOTAL_ROW, column=S_COL_TOTAL,
            value=f"=SUM(D{SUMMARY_FIRST_ROW}:D{SUMMARY_LAST_ROW})")
    for c in range(1, 5):
        cell = ws.cell(row=TOTAL_ROW, column=c)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        if c in (S_COL_COUNT, S_COL_TOTAL):
            cell.alignment = CENTER

    ws.cell(row=LOG_TITLE_ROW, column=1,
            value="ثبت تک‌تک فروش‌ها (هر ردیف = یک فروش، با قیمت واقعی همان لحظه)").font = TITLE_FONT

    log_headers = ["ردیف", "ساعت", "کلید", "نام نوشیدنی", "قیمت (تومان)"]
    for col, text in enumerate(log_headers, start=1):
        cell = ws.cell(row=LOG_HEADER_ROW, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    widths = {1: 16, 2: 24, 3: 16, 4: 24, 5: 18}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    return ws


def _load_workbook():
    _ensure_dir()
    if os.path.exists(WORKBOOK_PATH):
        return load_workbook(WORKBOOK_PATH)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _get_sheet_for_date(wb, date_str):
    if date_str in wb.sheetnames:
        return wb[date_str]
    return _new_sheet(wb, date_str)


def _get_today_sheet(wb):
    return _get_sheet_for_date(wb, _today_str())


def _first_empty_log_row(ws):
    row = LOG_FIRST_ROW
    while ws.cell(row=row, column=L_COL_NO).value is not None:
        row += 1
    return row


def _count_for_key(ws, item_key):
    count = 0
    row = LOG_FIRST_ROW
    while ws.cell(row=row, column=L_COL_NO).value is not None:
        if ws.cell(row=row, column=L_COL_KEY).value == item_key:
            count += 1
        row += 1
    return count


def _item_name(item_key):
    for item in MENU_ITEMS:
        if item["key"] == item_key:
            return item["name"]
    return item_key


def record_sale(item_key, price, date_str=None):
    """Append one sale row (with its own price - may differ from other
    sales of the same item that day) to the log of the given day (default:
    today). Use date_str to add a sale you forgot to enter earlier -
    returns the new total count for that item on that day."""
    wb = _load_workbook()
    ws = _get_sheet_for_date(wb, date_str or _today_str())

    row = _first_empty_log_row(ws)
    ws.cell(row=row, column=L_COL_NO, value=row - LOG_FIRST_ROW + 1).border = THIN_BORDER
    ws.cell(row=row, column=L_COL_TIME, value=datetime.now().strftime("%H:%M:%S")).border = THIN_BORDER
    ws.cell(row=row, column=L_COL_KEY, value=item_key).border = THIN_BORDER
    ws.cell(row=row, column=L_COL_NAME, value=_item_name(item_key)).border = THIN_BORDER
    ws.cell(row=row, column=L_COL_PRICE, value=price).border = THIN_BORDER

    wb.save(WORKBOOK_PATH)
    return _count_for_key(ws, item_key)


def decrement_sale(item_key, date_str=None):
    """Remove the most recent sale row for this item on the given day
    (default: today) - undo a misclick. Returns the new count."""
    wb = _load_workbook()
    ws = _get_sheet_for_date(wb, date_str or _today_str())

    last_row = None
    row = LOG_FIRST_ROW
    while ws.cell(row=row, column=L_COL_NO).value is not None:
        if ws.cell(row=row, column=L_COL_KEY).value == item_key:
            last_row = row
        row += 1

    if last_row is not None:
        ws.delete_rows(last_row, 1)
        # renumber the "ردیف" column so it stays sequential
        row = LOG_FIRST_ROW
        n = 1
        while ws.cell(row=row, column=L_COL_KEY).value is not None:
            ws.cell(row=row, column=L_COL_NO, value=n)
            n += 1
            row += 1
        wb.save(WORKBOOK_PATH)

    return _count_for_key(ws, item_key)


def _scan_log(ws):
    """Yield (item_key, price) for every sale row present in a day's sheet."""
    row = LOG_FIRST_ROW
    while ws.cell(row=row, column=L_COL_NO).value is not None:
        key = ws.cell(row=row, column=L_COL_KEY).value
        price = ws.cell(row=row, column=L_COL_PRICE).value or 0
        if key:
            yield key, price
        row += 1


def get_today_counts():
    """Return {item_key: count} for today by scanning the log. Empty dict
    if no file / no sheet for today yet."""
    if not os.path.exists(WORKBOOK_PATH):
        return {}
    wb = load_workbook(WORKBOOK_PATH)
    name = _today_str()
    if name not in wb.sheetnames:
        return {}
    ws = wb[name]
    counts = {item["key"]: 0 for item in MENU_ITEMS}
    for key, _price in _scan_log(ws):
        if key in counts:
            counts[key] += 1
    return counts


def list_available_dates():
    """All Jalali date strings (YYYY-MM-DD) that currently have a sheet,
    sorted from oldest to newest. Empty list if the workbook doesn't
    exist yet."""
    if not os.path.exists(WORKBOOK_PATH):
        return []
    wb = load_workbook(WORKBOOK_PATH, read_only=True)
    dates = [name for name in wb.sheetnames if date_utils.is_valid(name)]
    return sorted(dates)


def get_report(date_from, date_to):
    """Aggregate sales for every day between date_from and date_to
    (inclusive, both 'YYYY-MM-DD' Jalali strings). Returns:
        {
          "items": [{"key", "name", "count", "total"}, ...],  # menu order
          "total_count": int,
          "total_amount": number,
          "days_included": [list of date strings actually found],
        }
    """
    items_agg = {item["key"]: {"key": item["key"], "name": item["name"],
                                "count": 0, "total": 0} for item in MENU_ITEMS}
    days_included = []

    if os.path.exists(WORKBOOK_PATH):
        wb = load_workbook(WORKBOOK_PATH, read_only=True)
        for sheet_name in wb.sheetnames:
            if not date_utils.is_valid(sheet_name):
                continue
            if not (date_from <= sheet_name <= date_to):
                continue
            days_included.append(sheet_name)
            ws = wb[sheet_name]
            for key, price in _scan_log(ws):
                if key in items_agg:
                    items_agg[key]["count"] += 1
                    items_agg[key]["total"] += price

    items = [items_agg[item["key"]] for item in MENU_ITEMS]
    total_count = sum(i["count"] for i in items)
    total_amount = sum(i["total"] for i in items)
    return {
        "items": items,
        "total_count": total_count,
        "total_amount": total_amount,
        "days_included": sorted(days_included),
    }


def export_report(date_from, date_to, path):
    """Save the aggregated report for the given date range as a standalone
    .xlsx file at `path`."""
    report = get_report(date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش فروش"
    ws.sheet_view.rightToLeft = True

    title = f"گزارش فروش از {date_from} تا {date_to}"
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT

    headers = ["نام نوشیدنی", "تعداد فروش", "مبلغ کل (تومان)"]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    r = 3
    for item in report["items"]:
        ws.cell(row=r, column=1, value=item["name"]).border = THIN_BORDER
        c_cell = ws.cell(row=r, column=2, value=item["count"])
        t_cell = ws.cell(row=r, column=3, value=item["total"])
        c_cell.border = THIN_BORDER
        t_cell.border = THIN_BORDER
        c_cell.alignment = CENTER
        t_cell.alignment = CENTER
        r += 1

    ws.cell(row=r, column=1, value="مجموع کل")
    tc = ws.cell(row=r, column=2, value=report["total_count"])
    ta = ws.cell(row=r, column=3, value=report["total_amount"])
    for cell in (ws.cell(row=r, column=1), tc, ta):
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
    tc.alignment = CENTER
    ta.alignment = CENTER

    widths = {1: 26, 2: 16, 3: 18}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(path)
    return path
